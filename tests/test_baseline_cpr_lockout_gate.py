"""Lockout-gate tests for `baseline_CPR_pct` (issue #22).

The workbook gates `predicted_SMM` and every `CPR_attr_*` to 0 when a
loan is in lockout, but `baseline_CPR_pct` (Loan_Input column BW) is
currently the intercept-only constant ~1.645 for every loan, including
locked ones. That asymmetry produces a per-loan reconciliation gap on
`Deal_Comparison` (visible on `GNR_2024_NEW`, which has locked loans).

These tests verify the gate matches the one already on the SMM /
predicted-CPR path, in both the Python reference scorer
(`predict_one`) and the Excel formula (`write_loan_input`).
"""
import openpyxl
import pandas as pd
import pytest

from model.build_excel import write_loan_input
from model.predict_python import predict_one


# A locked loan from the shipped demo sample. The first row of
# `model/sample_loans.csv` has `lockout_end_date = 20610101` and
# `period = 202603`, so `in_lockout = 1`. The non-locked counterpart
# has `lockout_end_date` blank or already in the past.
def _locked_loan_row() -> pd.Series:
    row = pd.Series({
        "deal_id":           "GNR_2024_NEW",
        "period":            202603,
        "loan_rate":         5.69,
        "loan_rate_pct":     5.69,
        "upb":               24_499_231.76,
        "origination_date":  20251201,
        "loan_maturity_date": 20610101,
        "lockout_end_date":  20360201,  # well past period → in_lockout=1
        "prepay_end_date":   20360201,
        "prepay_premium_period_yrs": 10,
        "fha_program_code":  "232/223(F)",
        "affordable_status": "",
        "pool_type":         "PN",
        "vintage_year":      2025,
        "loan_age_months":   3.0,
        "sato_bps":          -13.0,
        "plc_rate_bps":      490.5,
        "prepay_penalty_points": 10.0,
        "refi_incentive_bps": -59.0,
    })
    return row


def _unlocked_loan_row() -> pd.Series:
    row = _locked_loan_row().copy()
    # Push lockout end before the period → in_lockout=0
    row["lockout_end_date"] = 20240101
    row["prepay_end_date"]  = 20240101
    return row


def test_predict_one_baseline_is_zero_for_locked_loan():
    """Acceptance criterion (issue #22): a loan in lockout contributes
    0 to the weighted baseline CPR, matching the existing gate on the
    SMM / predicted-CPR path. Today it contributes ~1.645 — which is
    exactly the reconciliation gap on `Deal_Comparison` for any deal
    with locked loans (most clearly `GNR_2024_NEW`)."""
    out = predict_one(_locked_loan_row())
    assert out["in_lockout"] == 1, "test fixture should be in lockout"
    assert out["predicted_CPR_pct"] == 0.0, (
        "fixture sanity: predicted CPR is already lockout-gated to 0"
    )
    assert out["baseline_CPR_pct"] == 0.0, (
        f"baseline_CPR_pct for a locked loan is "
        f"{out['baseline_CPR_pct']:.4f}; should be 0 to match the "
        f"lockout gate on predicted_SMM / predicted_CPR_pct"
    )


def test_predict_one_baseline_is_unchanged_for_unlocked_loan():
    """Non-regression: gating locked loans must not change the
    intercept-only baseline for non-locked loans. The level (~1.645%
    on the shipped coefficients) is the headline number ADR-0002
    reports across all six demo deals."""
    out = predict_one(_unlocked_loan_row())
    assert out["in_lockout"] == 0, "test fixture should NOT be in lockout"
    assert out["baseline_CPR_pct"] == pytest.approx(1.6454859537488487, rel=1e-9), (
        f"baseline_CPR_pct for a non-locked loan is "
        f"{out['baseline_CPR_pct']:.6f}; should still be the "
        f"intercept-only annualized CPR (~1.6455)"
    )


def _column_letter_for_header(ws, header: str) -> str:
    """Find the Loan_Input column letter whose row-2 header is `header`."""
    for cell in ws[2]:
        if cell.value == header:
            return cell.column_letter
    raise AssertionError(f"header {header!r} not found in Loan_Input row 2")


def test_workbook_baseline_cpr_formula_carries_lockout_gate():
    """Acceptance criterion (issue #22): the `baseline_CPR_pct` cell in
    the Loan_Input sheet must be lockout-gated, so a locked loan shows
    0 instead of the intercept-only 1.645. This mirrors the same
    `IF(in_lockout=1, 0, …)` guard that already wraps `predicted_SMM`
    and every `CPR_attr_*` column, and closes the per-loan
    reconciliation gap on `Deal_Comparison` (most visible on
    `GNR_2024_NEW`)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan_Input"
    write_loan_input(ws)

    baseline_col = _column_letter_for_header(ws, "baseline_CPR_pct")
    in_lockout_col = _column_letter_for_header(ws, "in_lockout")

    formula = ws[f"{baseline_col}3"].value
    assert isinstance(formula, str) and formula.startswith("="), (
        f"baseline_CPR_pct cell at {baseline_col}3 should be a formula; "
        f"got {formula!r}"
    )

    expected_gate = f"IF({in_lockout_col}3=1,0,"
    assert expected_gate in formula, (
        f"baseline_CPR_pct formula at {baseline_col}3 is missing the "
        f"lockout gate {expected_gate!r}. Got: {formula!r}"
    )
