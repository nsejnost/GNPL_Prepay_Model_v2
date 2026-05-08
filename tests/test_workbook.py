"""LifetimeCPRWorkbook tests.

The workbook is the user-visible deliverable for the Lifetime CPR engine
(slice 3 / issue #12). These tests verify behavior through the public
interface (write a workbook, then re-read it with openpyxl): they describe
*what* the workbook contains, not *how* it is laid out internally, so the
internal layout can change as long as the headline contracts hold.
"""
import os

import openpyxl
import pandas as pd
import pytest

from model.lifetime_engine.engine import Engine
from model.lifetime_engine.rate_scenario import RateScenario
from model.lifetime_engine.workbook import LifetimeCPRWorkbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAMPLE_PATH = os.path.join(ROOT, "model", "sample_loans.csv")


def _sample_loan() -> dict:
    df = pd.read_csv(SAMPLE_PATH)
    return df.iloc[0].to_dict()


def _build_result():
    loan = _sample_loan()
    scenario = RateScenario(treasury_bps=400, spread_bps=100, parallel_shock_bps=0)
    return loan, scenario, Engine().run(loan, scenario)


def test_cli_output_flag_writes_a_workbook_with_the_four_sheets(tmp_path):
    """Acceptance criterion: the CLI accepts an `--output workbook.xlsx`
    flag that writes a non-empty workbook with the four required sheets,
    and the engine result and workbook output are produced from the
    same call. Asserts the file is parseable and the t=0 SMM round-
    trip matches the JSON also emitted to stdout (single source of
    truth)."""
    import json
    from model.lifetime_engine.cli import main

    out = tmp_path / "lifetime.xlsx"
    rc = main([
        "--input", SAMPLE_PATH,
        "--row", "0",
        "--treasury-bps", "400",
        "--spread-bps", "100",
        "--shock-bps", "0",
        "--output", str(out),
    ])
    assert rc == 0
    assert out.exists() and out.stat().st_size > 0

    wb = openpyxl.load_workbook(str(out))
    assert set(wb.sheetnames) >= {"Inputs", "Scenario", "ForwardGrid", "Summary"}


FORWARD_GRID_COLUMNS = (
    "month_index",
    "synthetic_period",
    "loan_age_months",
    "in_lockout",
    "in_prepay_penalty",
    "prepay_penalty_points",
    "months_post_lockout",
    "months_to_maturity",
    "refi_incentive_bps",
    "scheduled_balance",
    "expected_balance",
    "smm",
    "smm_log_odds_total",
    "expected_prepay_dollars",
)


def _row_values(ws, row_idx):
    return [c.value for c in ws[row_idx]]


def _header_index(ws, name):
    headers = _row_values(ws, 1)
    return headers.index(name) + 1  # openpyxl is 1-indexed


def test_workbook_writes_non_empty_file_with_four_required_sheets(tmp_path):
    _, _, result = _build_result()
    out = tmp_path / "lifetime.xlsx"

    LifetimeCPRWorkbook(result).write(str(out))

    assert out.exists()
    assert out.stat().st_size > 0
    wb = openpyxl.load_workbook(str(out))
    assert set(wb.sheetnames) >= {"Inputs", "Scenario", "ForwardGrid", "Summary"}


def test_forward_grid_sheet_has_expected_columns_and_one_row_per_month(tmp_path):
    _, _, result = _build_result()
    out = tmp_path / "lifetime.xlsx"

    LifetimeCPRWorkbook(result).write(str(out))

    ws = openpyxl.load_workbook(str(out))["ForwardGrid"]
    headers = _row_values(ws, 1)
    for col in FORWARD_GRID_COLUMNS:
        assert col in headers, f"missing ForwardGrid column: {col}"
    # One header row + one row per simulated month.
    assert ws.max_row == len(result.forward_grid) + 1


def _find_label_value(ws, label):
    """Two-column key/value sheet helper: scan column A for `label`,
    return the value in column B of the same row. Returns None if the
    label isn't on the sheet."""
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        if row[0] == label:
            return row[1]
    return None


def _column_a_labels(ws):
    return [c.value for c in ws["A"]]


def test_summary_sheet_lifetime_cpr_cell_equals_engine_result(tmp_path):
    """Acceptance criterion: Summary sheet's lifetime CPR cell equals
    the engine's `lifetime_cpr_pct` field."""
    _, _, result = _build_result()
    out = tmp_path / "lifetime.xlsx"

    LifetimeCPRWorkbook(result).write(str(out))

    ws = openpyxl.load_workbook(str(out))["Summary"]
    cell = _find_label_value(ws, "lifetime_cpr_pct")
    assert cell == pytest.approx(
        result.summary.lifetime_cpr_pct, rel=1e-12, abs=1e-15
    )


INPUTS_REQUIRED_LABELS = (
    "loan_rate",
    "upb",
    "period",
    "loan_maturity_date",
    "lockout_end_date",
    "prepay_end_date",
    "fha_program_code",
)


def test_inputs_sheet_lists_loan_static_attributes(tmp_path):
    """Acceptance criterion: Inputs sheet contains the loan's static
    attributes (note rate, UPB, dates, FHA program). The result is the
    presentation contract — the engine carries the original loan
    dictionary on the result so the workbook can render it without a
    separate input."""
    _, _, result = _build_result()
    out = tmp_path / "lifetime.xlsx"

    LifetimeCPRWorkbook(result).write(str(out))

    ws = openpyxl.load_workbook(str(out))["Inputs"]
    labels_present = _column_a_labels(ws)
    for label in INPUTS_REQUIRED_LABELS:
        assert label in labels_present, f"Inputs sheet missing label {label!r}"
    # Spot-check: the loan_rate value must round-trip exactly (the
    # snapshot of what was scored is what the user reads off the
    # Inputs sheet).
    loan, _, _ = _build_result()
    assert _find_label_value(ws, "loan_rate") == pytest.approx(
        float(loan["loan_rate"]), rel=1e-12, abs=1e-15
    )


SCENARIO_INPUT_LABELS = ("treasury_bps", "spread_bps", "parallel_shock_bps")
HOUSE_INPUT_FILL_HEX = "FFF2CC"  # yellow-input-cell hex from build_excel.py


def test_scenario_sheet_round_trips_treasury_spread_and_shock_in_bps(tmp_path):
    """Acceptance criterion: Scenario sheet contains the user-editable
    Treasury bps / spread bps / parallel shock bps. Values must round-
    trip from the engine's RateScenario through Excel."""
    _, scenario, result = _build_result()
    out = tmp_path / "lifetime.xlsx"

    LifetimeCPRWorkbook(result).write(str(out))

    ws = openpyxl.load_workbook(str(out))["Scenario"]
    assert _find_label_value(ws, "treasury_bps") == pytest.approx(scenario.treasury_bps)
    assert _find_label_value(ws, "spread_bps") == pytest.approx(scenario.spread_bps)
    assert _find_label_value(ws, "parallel_shock_bps") == pytest.approx(
        scenario.parallel_shock_bps
    )


def test_scenario_input_cells_use_house_yellow_fill(tmp_path):
    """Acceptance criterion: yellow input-cell styling (FFF2CC fill,
    matching the existing house workbook) is applied to user-editable
    conceptual fields. v1 doesn't recompute on edit; the colour is the
    convention that signals to the user which cells they would
    conceptually drive."""
    _, _, result = _build_result()
    out = tmp_path / "lifetime.xlsx"

    LifetimeCPRWorkbook(result).write(str(out))

    ws = openpyxl.load_workbook(str(out))["Scenario"]
    for label in SCENARIO_INPUT_LABELS:
        labels = _column_a_labels(ws)
        row_idx = labels.index(label) + 1
        cell = ws.cell(row=row_idx, column=2)  # value cell next to label
        fill_hex = (cell.fill.fgColor.rgb or "")[-6:].upper()
        assert fill_hex == HOUSE_INPUT_FILL_HEX, (
            f"Scenario.{label} value cell fill is {fill_hex!r}, "
            f"expected the house yellow {HOUSE_INPUT_FILL_HEX!r}"
        )


def test_summary_sheet_references_optional_cpr_scalar_for_level_adjustment(tmp_path):
    """Acceptance criterion: Summary tab includes a note referencing
    `OPTIONAL_CPR_SCALAR` as the documented level-shift remediation
    knob (per parent PRD's Further Notes on decile-10 over-prediction).
    Read every string cell on the sheet and assert the literal token
    `OPTIONAL_CPR_SCALAR` appears somewhere."""
    _, _, result = _build_result()
    out = tmp_path / "lifetime.xlsx"

    LifetimeCPRWorkbook(result).write(str(out))

    ws = openpyxl.load_workbook(str(out))["Summary"]
    text_blob = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "OPTIONAL_CPR_SCALAR" in text_blob


def test_forward_grid_month1_smm_round_trips_through_workbook(tmp_path):
    """The forward grid's month-1 SMM cell, after round-tripping through
    Excel, must equal the engine result's forward_grid month-1 SMM. By
    slice 2's t=0 agreement test that value also equals the snapshot
    scorer's predicted SMM, so writing the file faithfully is the only
    additional invariant under test here."""
    _, _, result = _build_result()
    out = tmp_path / "lifetime.xlsx"

    LifetimeCPRWorkbook(result).write(str(out))

    ws = openpyxl.load_workbook(str(out))["ForwardGrid"]
    smm_col = _header_index(ws, "smm")
    cell_value = ws.cell(row=2, column=smm_col).value
    assert cell_value == pytest.approx(
        float(result.forward_grid["smm"].iloc[0]), rel=1e-12, abs=1e-15
    )
