"""Deal_Comparison row-label tests for the main GNPL workbook builder.

These tests target the public interface `model.build_excel.write_deal_comparison`
which writes the `Deal_Comparison` sheet of `GNPL_Prepay_Model.xlsx`. The
behavior under test is the row-27 label drift documented in issue #21
and ADR-0002: the row pulls `wtd_attr_POOL` from `Deal_Summary!AA` (which
in turn pulls `Loan_Input!CJ` = `CPR_attr_POOL`), but the label and driver
text on the sheet still say MOD / "Modified / non-level / mature". The
shipped model has a POOL component (pool-type dummies LM/PN/LS/RX), not a
separate modified/non-level/mature component.
"""
import openpyxl

from model.build_excel import write_deal_comparison


def _build_deal_comparison_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deal_Comparison"
    write_deal_comparison(ws)
    return ws


def test_row_27_label_reflects_pool_component_not_mod():
    """Issue #21 acceptance criterion: the row that carries the POOL
    component (pool-type LM/PN/LS/RX dummies) must be labeled as POOL on
    `Deal_Comparison`. Today it is mislabeled "CPR attr: MOD"."""
    ws = _build_deal_comparison_sheet()

    label = ws.cell(row=27, column=1).value
    assert label is not None, "row 27 should have a label in column A"
    assert "POOL" in label.upper(), (
        f"row 27 label is {label!r}; should reflect the POOL component "
        f"(pool-type LM/PN/LS/RX dummies) — see issue #21 / ADR-0002"
    )
    assert "MOD" not in label.upper().split(":")[-1].strip().split(), (
        f"row 27 label is {label!r}; should no longer reference MOD"
    )


def test_row_27_driver_text_describes_pool_type_dummies():
    """The driver-text column (E) explains what the row's underlying
    feature is. Row 27 carries the pool-type dummy component, so the
    explanation must reference pool type — not modified/non-level/mature
    flags, which are not a component in the shipped model."""
    ws = _build_deal_comparison_sheet()

    driver_text = ws.cell(row=27, column=5).value or ""
    assert "pool" in driver_text.lower(), (
        f"row 27 driver text is {driver_text!r}; should describe the "
        f"pool-type dummies (LM/PN/LS/RX) that this row actually carries"
    )
    assert "modified" not in driver_text.lower(), (
        f"row 27 driver text is {driver_text!r}; should no longer "
        f"describe the row as 'modified / non-level / mature'"
    )
