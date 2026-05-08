"""LifetimeCPRWorkbook: render a LifetimeCPRResult to an Excel workbook.

The output mirrors the existing house style from `model/build_excel.py`:
yellow input cells for user-editable conceptual fields, dark-blue header
rows, and a Summary sheet that documents the `OPTIONAL_CPR_SCALAR`
level-shift remediation knob.

Sheets:
  - Inputs       — the loan's static attributes
  - Scenario     — Treasury / spread / shock bps (yellow input cells)
  - ForwardGrid  — one row per simulated month
  - Summary      — lifetime SMM, lifetime CPR, peak SMM/month, total prepay
"""
from __future__ import annotations

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .engine import LifetimeCPRResult

# House style — values copied from `model/build_excel.py` so the workbook
# reads visually as part of the same deliverable family.
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")    # yellow = user input
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")    # dark blue header
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center")

FORWARD_GRID_COLUMNS = (
    "month_index",
    "synthetic_period",
    "loan_age_months",
    "in_lockout",
    "in_prepay_penalty",
    "prepay_penalty_points",
    "months_post_lockout",
    "months_to_maturity",
    "refi_incentive_bps",
    "scheduled_balance",
    "expected_balance",
    "smm",
    "smm_log_odds_total",
    "expected_prepay_dollars",
)


class LifetimeCPRWorkbook:
    def __init__(self, result: LifetimeCPRResult):
        self.result = result

    def write(self, output_path: str) -> None:
        wb = openpyxl.Workbook()
        inputs = wb.active
        inputs.title = "Inputs"
        self._write_inputs(inputs)
        self._write_scenario(wb.create_sheet("Scenario"))
        self._write_forward_grid(wb.create_sheet("ForwardGrid"))
        self._write_summary(wb.create_sheet("Summary"))
        wb.save(output_path)

    def _write_scenario(self, ws) -> None:
        scenario = self.result.scenario
        rows = (
            ("treasury_bps", scenario.treasury_bps),
            ("spread_bps", scenario.spread_bps),
            ("parallel_shock_bps", scenario.parallel_shock_bps),
        )
        _write_key_value_sheet(ws, "Field", "Value", rows, value_fill=INPUT_FILL)

    def _write_inputs(self, ws) -> None:
        labels = (
            "loan_rate", "upb", "period",
            "origination_date", "loan_maturity_date",
            "lockout_end_date", "prepay_end_date",
            "prepay_premium_period_yrs",
            "fha_program_code", "pool_type",
            "loan_purpose", "affordable_status",
            "sato_bps", "modified", "non_level", "mature",
        )
        loan = self.result.loan_inputs
        rows = ((label, loan.get(label)) for label in labels)
        _write_key_value_sheet(ws, "Field", "Value", rows)

    def _write_forward_grid(self, ws) -> None:
        for col_idx, col in enumerate(FORWARD_GRID_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        grid = self.result.forward_grid
        for row_idx, (_, row) in enumerate(grid.iterrows(), start=2):
            for col_idx, col in enumerate(FORWARD_GRID_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=_to_excel(row[col]))
        for col_idx in range(1, len(FORWARD_GRID_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    def _write_summary(self, ws) -> None:
        s = self.result.summary
        headline_rows = (
            ("lifetime_smm", s.lifetime_smm),
            ("lifetime_cpr_pct", s.lifetime_cpr_pct),
            ("baseline_lifetime_cpr_pct", self.result.baseline_lifetime_cpr_pct),
            ("peak_smm", s.peak_smm),
            ("peak_smm_month", s.peak_smm_month),
            ("expected_total_prepay", s.expected_total_prepay),
        )
        driver_rows = tuple(
            sorted(
                self.result.driver_attribution.items(),
                key=lambda item: -abs(item[1]),
            )
        )
        rows = headline_rows + driver_rows
        _write_key_value_sheet(ws, "Field", "Value", rows)
        note_row = len(rows) + 3
        ws.cell(
            row=note_row, column=1,
            value=(
                "Level adjustment: the parent model is documented to over-predict "
                "in steeply out-of-the-money rate regimes (see Further Notes on "
                "decile-10 over-prediction in the parent PRD). To level-shift the "
                "predicted CPR for a stress or calibration scenario, multiply the "
                "engine's SMM path by the OPTIONAL_CPR_SCALAR cell (Parameters "
                "sheet of the house workbook); 1.0 is the unbiased default."
            ),
        )


def _to_excel(value):
    """openpyxl rejects numpy scalars; cast to native Python types."""
    if isinstance(value, np.generic):
        return value.item()
    return value


def _write_key_value_sheet(ws, key_header, value_header, rows, value_fill=None):
    """Write a two-column key/value sheet with the house header style."""
    header_a = ws.cell(row=1, column=1, value=key_header)
    header_b = ws.cell(row=1, column=2, value=value_header)
    for c in (header_a, header_b):
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        cell = ws.cell(row=i, column=2, value=_to_excel(value))
        if value_fill is not None:
            cell.fill = value_fill
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
