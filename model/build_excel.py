"""
GNMA Project Loan Prepayment Model — Excel Builder
====================================================

Generates ``GNPL_Prepay_Model.xlsx``, a fully-formula-driven workbook
that lets a user paste loan-level data for any number of GNR REMIC
deals and get back loan-level CPR predictions, per-feature attribution,
and deal-level aggregation/comparison.

The workbook is "live": every prediction is an Excel formula reading
from the Parameters sheet, so a user can recalibrate by editing
coefficients without re-running this builder.
"""

from __future__ import annotations

import json
import os

import openpyxl
import pandas as pd
from openpyxl.styles import (Alignment, Border, Font, NumberFormatDescriptor,
                              PatternFill, Side)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MODEL_DIR = os.path.join(ROOT, "model")
OUT_XLSX  = os.path.join(ROOT, "GNPL_Prepay_Model.xlsx")

with open(os.path.join(MODEL_DIR, "feature_metadata.json")) as f:
    META = json.load(f)
COEF = pd.read_csv(os.path.join(MODEL_DIR, "coefficients.csv"))
SAMPLE = pd.read_csv(os.path.join(MODEL_DIR, "sample_loans.csv"))
PLC_HIST = pd.read_csv(os.path.join(MODEL_DIR, "plc_rates_history.csv"))
VINTAGE  = pd.read_csv(os.path.join(MODEL_DIR, "vintage_median_rates.csv"))

INTERCEPT = float(COEF.loc[COEF["feature"] == "__intercept__",
                           "coefficient"].iloc[0])
COEF_NO_INT = COEF[COEF["feature"] != "__intercept__"].copy()

# ─── Styles ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL  = PatternFill("solid", fgColor="FFF2CC")  # yellow = user input
DERIVED_FILL = PatternFill("solid", fgColor="E2EFDA")  # green = formula
ATTR_FILL   = PatternFill("solid", fgColor="FCE4D6")  # orange = attribution
TOTAL_FILL  = PatternFill("solid", fgColor="C6E0B4")
PARAM_FILL  = PatternFill("solid", fgColor="DDEBF7")

HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def autosize(ws, min_w: int = 10, max_w: int = 36) -> None:
    """Auto-size columns based on content; bounded so wide columns
    don't blow up the screen real-estate.  Skips merged cells which
    don't expose column_letter."""
    n_cols = ws.max_column
    n_rows = ws.max_row
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, n_rows + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            v = cell.value
            if v is None:
                continue
            try:
                length = len(str(v))
            except Exception:
                length = 10
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = max(min_w,
                                                     min(max_w, max_len + 2))


# ─── Build workbook ─────────────────────────────────────────────────

def build():
    wb = openpyxl.Workbook()
    # Set the workbook's default sheet — we'll rename it to README
    readme = wb.active
    readme.title = "README"

    write_readme(readme)
    params = wb.create_sheet("Parameters")
    write_parameters(params)
    plc_ws = wb.create_sheet("PLC_Rates")
    write_plc_rates(plc_ws)
    vint_ws = wb.create_sheet("Vintage_Rates")
    write_vintage_rates(vint_ws)
    fha_ws = wb.create_sheet("FHA_Lookup")
    write_fha_lookup(fha_ws)
    loan_ws = wb.create_sheet("Loan_Input")
    n_loans = write_loan_input(loan_ws)
    deal_ws = wb.create_sheet("Deal_Summary")
    write_deal_summary(deal_ws, loan_ws, n_loans)
    cmp_ws = wb.create_sheet("Deal_Comparison")
    write_deal_comparison(cmp_ws)
    val_ws = wb.create_sheet("Validation")
    write_validation(val_ws)

    define_named_ranges(wb)
    wb.save(OUT_XLSX)
    print(f"[save] {OUT_XLSX}")
    print(f"       {n_loans} sample loans pre-loaded across "
          f"{SAMPLE['deal_id'].nunique()} demo deals")


# ─── README ─────────────────────────────────────────────────────────

README_LINES = [
    ("GNMA Project Loan Prepayment Model", BOLD, 16),
    ("", None, None),
    ("PURPOSE", BOLD, 12),
    ("This workbook predicts monthly SMM (Single Monthly Mortality) and "
     "annualized CPR (Conditional Prepayment Rate) for individual GNMA "
     "multifamily project loans, then aggregates loans up to the GNR REMIC "
     "deal level so deals can be compared and the drivers of speed "
     "differences can be isolated.", None, None),
    ("", None, None),
    ("HOW TO USE", BOLD, 12),
    ("1.  Open the Loan_Input sheet.", None, None),
    ("2.  Replace the demo data (rows 2 and below) with your own loan-level "
     "data.  At minimum you need: deal_id, loan_rate, upb, dates "
     "(origination, maturity, lockout end, prepay end), prepay penalty "
     "period, FHA program code, affordable status, modified/non-level/mature "
     "flags, period (YYYYMM).  All formulas to the right derive features "
     "and predict CPR automatically.", None, None),
    ("3.  Yellow cells are user input.  Green cells are derived/computed.  "
     "Orange cells are per-component attribution.", None, None),
    ("4.  The Deal_Summary sheet aggregates loans by deal_id with weighted "
     "averages.  The Deal_Comparison sheet displays the underlying drivers "
     "of CPR differences across deals side-by-side.", None, None),
    ("5.  To override a model coefficient (e.g. for a stress scenario), "
     "edit the corresponding cell on Parameters.", None, None),
    ("", None, None),
    ("MODEL ARCHITECTURE", BOLD, 12),
    (f"Trained on {META['train_n']:,} loan-month observations from "
     f"{META['train_period_min']} through {META['train_period_max']} "
     f"({META['train_events']:,} voluntary prepay events).  Held-out test: "
     f"{META['test_n']:,} observations from {META['test_period_min']} "
     f"through {META['test_period_max']}.", None, None),
    ("", None, None),
    ("Functional form: logistic regression on engineered piecewise-linear "
     "splines.  log( SMM/(1-SMM) ) = intercept + Σ contributions.  Each "
     "contribution maps cleanly to an Excel formula and to a CPR "
     "attribution.", None, None),
    ("", None, None),
    ("Components:", BOLD, 11),
    ("  •  Refi incentive S-curve (loan rate vs penalty-adjusted PLC).", None, None),
    ("  •  Loan age ramp + burnout (piecewise-linear seasoning).", None, None),
    ("  •  Penalty point S-curve (interaction with refi incentive).", None, None),
    ("  •  SATO (spread at origination — borrower quality proxy).", None, None),
    ("  •  Loan size effect (larger loans refi more easily).", None, None),
    ("  •  Months-to-maturity (balloon-driven refi pull).", None, None),
    ("  •  Months-post-lockout (pent-up demand).", None, None),
    ("  •  Refi × penalty interaction.", None, None),
    ("  •  In-penalty-window phase indicator.", None, None),
    ("  •  FHA program category dummies (221d4, 223a7, 232, 538, 241, 220, OTHER vs 223f).", None, None),
    ("  •  Loan purpose (NC vs RP).", None, None),
    ("  •  Affordable status (AFF/BAF/MKT vs unknown).", None, None),
    ("  •  Pool-type dummies (LM/PN/LS/RX vs reference).", None, None),
    ("", None, None),
    ("VALIDATION", BOLD, 12),
    (f"Train AUC: {META['train_auc']:.3f}, predicted CPR "
     f"{META['train_pred_cpr_pct']:.2f}% vs actual "
     f"{META['train_actual_cpr_pct']:.2f}%.", None, None),
    (f"Test  AUC: {META['test_auc']:.3f}, predicted CPR "
     f"{META['test_pred_cpr_pct']:.2f}% vs actual "
     f"{META['test_actual_cpr_pct']:.2f}%.", None, None),
    ("", None, None),
    ("The test period (Jul 2024 – Mar 2026) is a steeply out-of-the-money "
     "regime in which observed prepay rates are unusually low.  The model "
     "over-predicts in absolute level but ranks loans correctly.  This "
     "regime sensitivity is well-documented in the BAM and Citi project "
     "loan model papers (see /pdfs).  Use the OPTIONAL_CPR_SCALAR field on "
     "Parameters to apply a level adjustment if desired.", None, None),
    ("", None, None),
    ("KEY ASSUMPTIONS / LIMITATIONS", BOLD, 12),
    ("  •  Construction-phase loans (CL/CS pool types) are excluded — they "
     "are 0% voluntary prepay by definition.", None, None),
    ("  •  In-lockout loans are predicted as 0% voluntary prepay (capture "
     "scheduled by formula; manual override possible).", None, None),
    ("  •  This is a VOLUNTARY prepayment model only.  Involuntary buyouts "
     "(default-driven repurchase) are not predicted here — they are sparse "
     "(~0.02% per month) and require a credit model.", None, None),
    ("  •  Property-price-appreciation cashout dynamics are NOT explicitly "
     "modeled (no PPA inputs in GNMA disclosure).  They are captured "
     "implicitly via SATO and vintage effects.", None, None),
    ("  •  All dates use YYYYMMDD numeric format (e.g. 20240315 = "
     "2024-03-15).  Period uses YYYYMM (e.g. 202603 = March 2026).", None, None),
    ("", None, None),
    ("LEGEND", BOLD, 12),
    ("Yellow cells = user input", None, None),
    ("Green cells  = formula-derived feature", None, None),
    ("Orange cells = attribution to predicted CPR", None, None),
    ("Blue cells   = parameter / coefficient", None, None),
]


def write_readme(ws):
    for row_idx, (text, font, size) in enumerate(README_LINES, start=1):
        c = ws.cell(row=row_idx, column=1, value=text)
        if font is not None:
            c.font = Font(bold=True, size=size or 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    for r in range(1, len(README_LINES) + 1):
        ws.row_dimensions[r].height = max(15, 15 if not README_LINES[r-1][0]
                                          else min(75, 18 + len(README_LINES[r-1][0]) // 90 * 15))


# ─── Parameters sheet ───────────────────────────────────────────────

def write_parameters(ws):
    """Coefficient table.  Named ranges defined later for use in
    Loan_Input formulas."""
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["C1"] = "Description"
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    ws["A2"] = "INTERCEPT"
    ws["B2"] = INTERCEPT
    ws["C2"] = "Logistic intercept; baseline log-odds when all features = 0."
    ws["B2"].fill = PARAM_FILL
    ws["B2"].number_format = "0.00000000"

    ws["A3"] = "OPTIONAL_CPR_SCALAR"
    ws["B3"] = 1.00
    ws["C3"] = ("Multiplier on final predicted SMM. Use 1.0 for the unbiased "
                "model; use <1 to dampen if you believe the regime is more "
                "out-of-money than training; >1 for stress.")
    ws["B3"].fill = PARAM_FILL
    ws["B3"].number_format = "0.00"

    # Coefficients
    ws["A5"] = "Component"
    ws["B5"] = "Feature"
    ws["C5"] = "Coefficient"
    ws["D5"] = "Description"
    for c in ws[5]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    DESCRIPTIONS = {
        # refi splines
        "refi_lin":   "Linear slope of refi incentive in bps",
        "refi_k-150": "Hinge: extra slope past −150 bps incentive",
        "refi_k-50":  "Hinge: extra slope past −50 bps incentive",
        "refi_k0":    "Hinge: extra slope past 0 bps (at-money)",
        "refi_k50":   "Hinge: extra slope past +50 bps incentive",
        "refi_k100":  "Hinge: extra slope past +100 bps incentive",
        "refi_k150":  "Hinge: extra slope past +150 bps incentive",
        "refi_k250":  "Hinge: extra slope past +250 bps incentive",
        # age splines
        "age_lin":    "Linear slope of loan age (months)",
        "age_k12":    "Hinge: extra slope past 12 months age",
        "age_k36":    "Hinge: extra slope past 36 months age",
        "age_k60":    "Hinge: extra slope past 60 months age",
        "age_k84":    "Hinge: extra slope past 84 months age",
        "age_k120":   "Hinge: extra slope past 120 months age (burnout)",
        "age_k180":   "Hinge: extra slope past 180 months age",
        # pen splines
        "pen_lin":    "Linear slope of penalty points",
        "pen_k0":     "Hinge: extra slope past 0 pts",
        "pen_k2":     "Hinge: extra slope past 2 pts",
        "pen_k5":     "Hinge: extra slope past 5 pts",
        "pen_k8":     "Hinge: extra slope past 8 pts",
        # sato splines
        "sato_lin":   "Linear slope of SATO (bps)",
        "sato_k-100": "Hinge: extra slope past −100 bps SATO",
        "sato_k-25":  "Hinge: extra slope past −25 bps SATO",
        "sato_k25":   "Hinge: extra slope past +25 bps SATO",
        "sato_k100":  "Hinge: extra slope past +100 bps SATO",
        # size splines
        "size_lin":   "Linear slope of UPB ($M)",
        "size_k2":    "Hinge: extra slope past $2M UPB",
        "size_k10":   "Hinge: extra slope past $10M UPB",
        "size_k25":   "Hinge: extra slope past $25M UPB",
        # m2m splines
        "m2m_lin":    "Linear slope of months-to-maturity",
        "m2m_k12":    "Hinge: extra slope past 12 months to maturity",
        "m2m_k24":    "Hinge: extra slope past 24 months to maturity",
        "m2m_k60":    "Hinge: extra slope past 60 months to maturity",
        # mpl splines
        "mpl_lin":    "Linear slope of months-post-lockout",
        "mpl_k6":     "Hinge: extra slope past 6 months post-lockout",
        "mpl_k12":    "Hinge: extra slope past 12 months post-lockout",
        "mpl_k24":    "Hinge: extra slope past 24 months post-lockout",
        # interaction
        "refi_x_pen": "Refi×penalty interaction (refi_pos × pen_clip / 100)",
        # phase
        "in_penalty": "1 if in prepay-penalty window, else 0",
        # FHA
        "fha_221d4":  "FHA 221(d)(4) new-construction multifamily dummy",
        "fha_223a7":  "FHA 223(a)(7) streamlined refi dummy",
        "fha_232":    "FHA 232 healthcare dummy",
        "fha_538":    "USDA 538 rural multifamily dummy",
        "fha_241":    "FHA 241 supplemental dummy",
        "fha_220":    "FHA 220 urban renewal dummy",
        "fha_other":  "Other / uncategorized FHA programs",
        # purpose
        "is_nc":      "New construction (vs refi/purchase reference)",
        # affordable
        "aff_aff":    "Affordable status = AFF dummy",
        "aff_baf":    "Affordable status = BAF dummy",
        "aff_mkt":    "Affordable status = MKT dummy",
        # pool-type flags (live signal — legacy modified_ind/non_level_ind/
        # mature_loan_flag are blank in modern panel; pool_type carries
        # the modification/non-level/small-balance signal directly)
        "is_lm_pool": "Pool type = LM (Mature/Modified — IRR refi destination)",
        "is_pn_pool": "Pool type = PN (Non-Level amortization, dominant Project Loan pool)",
        "is_ls_pool": "Pool type = LS (Small-Balance Loan)",
        "is_rx_pool": "Pool type = RX (Mark-to-Market)",
    }

    def comp_of(name: str) -> str:
        if name.startswith("refi_x"):  return "INTERACTION"
        if name.startswith("refi"):    return "REFI"
        if name.startswith("age"):     return "AGE"
        if name.startswith("pen"):     return "PENALTY"
        if name.startswith("sato"):    return "SATO"
        if name.startswith("size"):    return "SIZE"
        if name.startswith("m2m"):     return "M2M"
        if name.startswith("mpl"):     return "MPL"
        if name.startswith("in_penalty"): return "PHASE"
        if name.startswith("fha"):     return "FHA"
        if name.startswith("is_nc"):   return "PURPOSE"
        if name.startswith("aff"):     return "AFFORDABLE"
        if name.startswith("is_"):     return "POOL"
        return "OTHER"

    row = 6
    for _, r in COEF_NO_INT.iterrows():
        ws.cell(row=row, column=1, value=comp_of(r["feature"]))
        ws.cell(row=row, column=2, value=r["feature"]).font = BOLD
        c = ws.cell(row=row, column=3, value=float(r["coefficient"]))
        c.number_format = "0.00000000"
        c.fill = PARAM_FILL
        ws.cell(row=row, column=4, value=DESCRIPTIONS.get(r["feature"], ""))
        row += 1

    autosize(ws, min_w=12, max_w=70)


# ─── PLC rate history (for refi incentive lookups) ──────────────────

def write_plc_rates(ws):
    ws["A1"] = "period"
    ws["B1"] = "PLC_rate_bps"
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, r in PLC_HIST.reset_index(drop=True).iterrows():
        ws.cell(row=i+2, column=1, value=int(r["period"]))
        ws.cell(row=i+2, column=2, value=float(r["plc_rate_bps"]))
    autosize(ws)


def write_vintage_rates(ws):
    ws["A1"] = "vintage_year"
    ws["B1"] = "median_loan_rate_pct"
    ws["C1"] = "n_loans"
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, r in VINTAGE.reset_index(drop=True).iterrows():
        ws.cell(row=i+2, column=1, value=int(r["vintage_year"]))
        ws.cell(row=i+2, column=2, value=float(r["median_rate"]))
        ws.cell(row=i+2, column=3, value=int(r["n_loans"]))
    autosize(ws)


def write_fha_lookup(ws):
    """Map FHA program code (free text) to category and loan_purpose.

    The lookup uses substring matching in priority order.  Excel can do
    this with nested IFERROR(SEARCH(...)) but the table makes the rules
    transparent and editable.
    """
    rules = [
        # (search_substring, fha_category, loan_purpose, is_nc_or_rp_note)
        ("538",    "538",    "538"  ),  # USDA rural — 538 priority
        ("232",    "232",    "NC"   ),  # standalone 232 (overridden by 223 below)
        ("223(F)", "223f",   "RP"   ),
        ("223F",   "223f",   "RP"   ),
        ("223(A)", "223a7",  "RP"   ),
        ("223A",   "223a7",  "RP"   ),
        ("221(D)", "221d4",  "NC"   ),
        ("221D",   "221d4",  "NC"   ),
        ("220",    "220",    "NC"   ),
        ("241",    "241",    "NC"   ),
    ]
    ws["A1"] = "search_substring"
    ws["B1"] = "fha_category"
    ws["C1"] = "loan_purpose"
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, (s, c1, c2) in enumerate(rules, start=2):
        ws.cell(row=i, column=1, value=s)
        ws.cell(row=i, column=2, value=c1)
        ws.cell(row=i, column=3, value=c2)
    autosize(ws)


# ─── Loan_Input sheet (the heavy lifting) ────────────────────────────

INPUT_COLUMNS = [
    # (header, source_field_or_None, is_input)
    ("deal_id",                    "deal_id",                    True),
    ("pool_cusip",                 "pool_cusip",                 True),
    ("pool_type",                  "pool_type",                  True),
    ("issuer_name",                "issuer_name",                True),
    ("loan_id",                    "loan_id",                    True),
    ("period",                     "period",                     True),
    ("loan_rate_pct",              "loan_rate",                  True),
    ("upb",                        "upb",                        True),
    ("loan_term_months",           "loan_term",                  True),
    ("origination_date",           "origination_date",           True),
    ("loan_maturity_date",         "loan_maturity_date",         True),
    ("lockout_end_date",           "lockout_end_date",           True),
    ("prepay_end_date",            "prepay_end_date",            True),
    ("prepay_premium_period_yrs",  "prepay_premium_period_yrs",  True),
    ("fha_program_code",           "fha_program_code",           True),
    ("affordable_status",          "affordable_status",          True),
    ("modified_ind",               "modified_ind",               True),
    ("non_level_ind",              "non_level_ind",              True),
    ("mature_loan_flag",           "mature_loan_flag",           True),
    ("property_state",             "property_state",             True),
    ("num_units",                  "num_units",                  True),
    # Pre-computed feature inputs (user can override; otherwise use the
    # ones derived in main.py / the parquet).  Including these as inputs
    # avoids requiring Excel to do prepay-schedule string parsing.
    ("plc_rate_bps_input",         "plc_rate_bps",               True),
    ("loan_age_months_input",      "loan_age_months",            True),
    ("vintage_year_input",         "vintage_year",               True),
    ("sato_bps_input",             "sato_bps",                   True),
    ("prepay_penalty_points_input","prepay_penalty_points",      True),
]

# Excel column letter for each
def col_letter(idx_1based: int) -> str:
    return get_column_letter(idx_1based)


def write_loan_input(ws) -> int:
    """Build the Loan_Input sheet.

    Returns the number of data rows written so the Deal_Summary sheet
    can size its formulas accordingly.
    """
    # Header row 1 — section labels (merged)
    # Header row 2 — actual column names
    # Data starts at row 3

    # Section blocks
    sections = [
        ("LOAN INPUTS (paste here)",          "USER",      INPUT_FILL),
        ("DERIVED FEATURES",                  "DERIVED",   DERIVED_FILL),
        ("LOG-ODDS CONTRIBUTIONS",            "LOGODDS",   ATTR_FILL),
        ("PREDICTION & ATTRIBUTION (CPR)",    "PRED",      TOTAL_FILL),
    ]

    # Build the column layout
    layout = []   # list of (header, group, formula_template_or_None)
    for header, _, _ in INPUT_COLUMNS:
        layout.append((header, "USER", None))

    # Derived feature columns (formulas reference INPUT columns)
    # Letters for inputs: deal_id=A, pool_cusip=B, ..., prepay_penalty_points_input=Z
    # We build a lookup of input header → column letter.
    input_header_to_col = {h: col_letter(i + 1)
                           for i, (h, _, _) in enumerate(INPUT_COLUMNS)}

    # We'll define the derived columns after the input block:
    derived = [
        ("loan_age_months",        # use input override else compute
         "USE_INPUT_OR_DERIVE",
         input_header_to_col["loan_age_months_input"],
         input_header_to_col["origination_date"],
         input_header_to_col["period"],
         "AGE",
         ),
        ("months_to_maturity",
         "M2M",
         input_header_to_col["loan_maturity_date"],
         input_header_to_col["period"],
         ),
        ("months_post_lockout",
         "MPL",
         input_header_to_col["lockout_end_date"],
         input_header_to_col["period"],
         ),
        ("in_lockout",
         "IN_LOCKOUT",
         input_header_to_col["lockout_end_date"],
         input_header_to_col["period"],
         ),
        ("in_prepay_penalty",
         "IN_PEN",
         input_header_to_col["prepay_end_date"],
         input_header_to_col["period"],
         ),
        ("past_all_restrictions",
         "PAST_ALL",
         ),
        ("vintage_year",
         "VINTAGE",
         input_header_to_col["vintage_year_input"],
         input_header_to_col["origination_date"],
         ),
        ("vintage_median_rate",
         "VMR",
         ),
        ("sato_bps",
         "SATO",
         input_header_to_col["sato_bps_input"],
         input_header_to_col["loan_rate_pct"],
         ),
        ("plc_rate_bps",
         "PLC",
         input_header_to_col["plc_rate_bps_input"],
         input_header_to_col["period"],
         ),
        ("prepay_penalty_points",
         "PEN",
         input_header_to_col["prepay_penalty_points_input"],
         input_header_to_col["prepay_end_date"],
         input_header_to_col["period"],
         input_header_to_col["prepay_premium_period_yrs"],
         ),
        ("gross_refi_incentive_bps",
         "GROSS_REFI",
         input_header_to_col["loan_rate_pct"],
         ),
        ("refi_incentive_bps",
         "REFI",
         input_header_to_col["loan_rate_pct"],
         ),
        ("upb_millions",
         "UPB_M",
         input_header_to_col["upb"],
         ),
        ("fha_category",
         "FHACAT",
         input_header_to_col["fha_program_code"],
         ),
        ("loan_purpose",
         "PURPOSE",
         input_header_to_col["fha_program_code"],
         ),
        ("is_nc",
         "ISNC",
         ),
        ("aff_aff",
         "AFFAFF",
         input_header_to_col["affordable_status"],
         ),
        ("aff_baf",
         "AFFBAF",
         input_header_to_col["affordable_status"],
         ),
        ("aff_mkt",
         "AFFMKT",
         input_header_to_col["affordable_status"],
         ),
        ("fha_221d4",
         "FHA221",
         ),
        ("fha_223a7",
         "FHA223A7",
         ),
        ("fha_232",
         "FHA232",
         ),
        ("fha_538",
         "FHA538",
         ),
        ("fha_241",
         "FHA241",
         ),
        ("fha_220",
         "FHA220",
         ),
        ("fha_other",
         "FHAOTHER",
         ),
        ("is_lm_pool",
         "ISLM",
         input_header_to_col["pool_type"],
         ),
        ("is_pn_pool",
         "ISPN",
         input_header_to_col["pool_type"],
         ),
        ("is_ls_pool",
         "ISLS",
         input_header_to_col["pool_type"],
         ),
        ("is_rx_pool",
         "ISRX",
         input_header_to_col["pool_type"],
         ),
    ]

    for tup in derived:
        layout.append((tup[0], "DERIVED", tup))

    # Component log-odds contributions (one column per component group)
    components = [
        "REFI_logodds",
        "AGE_logodds",
        "PEN_logodds",
        "SATO_logodds",
        "SIZE_logodds",
        "M2M_logodds",
        "MPL_logodds",
        "INTERACT_logodds",
        "PHASE_logodds",
        "FHA_logodds",
        "PURPOSE_logodds",
        "AFF_logodds",
        "POOL_logodds",
    ]
    for c in components:
        layout.append((c, "LOGODDS", None))

    # Final prediction & attribution columns
    pred_cols = [
        ("intercept_logodds", "PRED"),
        ("total_logodds", "PRED"),
        ("predicted_SMM", "PRED"),
        ("predicted_CPR_pct", "PRED"),
        ("baseline_CPR_pct", "PRED"),
        # Per-component CPR attribution: how many CPR percentage points
        # is each component adding (positive) or subtracting (negative)?
        ("CPR_attr_REFI", "PRED"),
        ("CPR_attr_AGE", "PRED"),
        ("CPR_attr_PEN", "PRED"),
        ("CPR_attr_SATO", "PRED"),
        ("CPR_attr_SIZE", "PRED"),
        ("CPR_attr_M2M", "PRED"),
        ("CPR_attr_MPL", "PRED"),
        ("CPR_attr_INTERACT", "PRED"),
        ("CPR_attr_PHASE", "PRED"),
        ("CPR_attr_FHA", "PRED"),
        ("CPR_attr_PURPOSE", "PRED"),
        ("CPR_attr_AFF", "PRED"),
        ("CPR_attr_POOL", "PRED"),
    ]
    for h, g in pred_cols:
        layout.append((h, g, None))

    # ── Write headers (rows 1 + 2) ──────────────────────────────────
    ws.cell(row=2, column=1)  # placeholder
    for i, (h, group, _) in enumerate(layout, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        if group == "USER":
            cell.fill = HEADER_FILL
        elif group == "DERIVED":
            cell.fill = PatternFill("solid", fgColor="375623")
        elif group == "LOGODDS":
            cell.fill = PatternFill("solid", fgColor="843C0C")
        else:
            cell.fill = PatternFill("solid", fgColor="2F5496")

    # Section banners (row 1)
    def section_span(group: str) -> tuple[int, int] | None:
        cols = [i for i, (_, g, _) in enumerate(layout, start=1) if g == group]
        if not cols:
            return None
        return (min(cols), max(cols))

    banners = [
        ("LOAN INPUTS — replace with your data",  "USER",     "FFD966"),
        ("DERIVED FEATURES — auto-computed",       "DERIVED",  "C6E0B4"),
        ("LOG-ODDS CONTRIBUTIONS — by component",  "LOGODDS",  "F4B084"),
        ("PREDICTED CPR & ATTRIBUTION",            "PRED",     "9DC3E6"),
    ]
    for label, group, color in banners:
        sp = section_span(group)
        if sp:
            start, end = sp
            ws.cell(row=1, column=start, value=label)
            ws.cell(row=1, column=start).font = Font(bold=True, color="000000")
            ws.cell(row=1, column=start).fill = PatternFill("solid", fgColor=color)
            ws.cell(row=1, column=start).alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=1, end_row=1,
                           start_column=start, end_column=end)

    # Pre-load sample rows
    n_rows = len(SAMPLE)
    for i in range(n_rows):
        s = SAMPLE.iloc[i]
        for col_idx, (h, group, _) in enumerate(layout, start=1):
            if group != "USER":
                continue
            src = next((sf for hdr, sf, _ in INPUT_COLUMNS if hdr == h), None)
            if src is None:
                continue
            v = s.get(src)
            if pd.isna(v):
                ws.cell(row=i + 3, column=col_idx, value=None)
            else:
                ws.cell(row=i + 3, column=col_idx, value=v)

    # ── Build the formula columns ───────────────────────────────────
    # Map header to column letter
    HC = {h: col_letter(i + 1) for i, (h, _, _) in enumerate(layout)}

    last_row = n_rows + 2  # data rows are 3 .. (n+2)

    for r in range(3, last_row + 1):

        def f(col: str, formula: str):
            ws.cell(row=r, column=ws[f"{col}2"].column,
                    value=formula)

        # ── DERIVED features ────────────────────────────────────────

        # loan_age_months: prefer override; else compute from origination_date and period.
        f(HC["loan_age_months"],
          f'=IF(ISNUMBER({HC["loan_age_months_input"]}{r}),'
          f'{HC["loan_age_months_input"]}{r},'
          f'IFERROR(MAX(0,'
          f'(VALUE(LEFT({HC["period"]}{r},4))-'
          f'VALUE(LEFT(TEXT({HC["origination_date"]}{r},"00000000"),4)))*12'
          f'+(VALUE(MID({HC["period"]}{r},5,2))-'
          f'VALUE(MID(TEXT({HC["origination_date"]}{r},"00000000"),5,2)))),""))')

        # months_to_maturity
        f(HC["months_to_maturity"],
          f'=IFERROR((VALUE(LEFT(TEXT({HC["loan_maturity_date"]}{r},"00000000"),4))-'
          f'VALUE(LEFT({HC["period"]}{r},4)))*12+'
          f'(VALUE(MID(TEXT({HC["loan_maturity_date"]}{r},"00000000"),5,2))-'
          f'VALUE(MID({HC["period"]}{r},5,2))),"")')

        # months_post_lockout (clamped 0..60)
        f(HC["months_post_lockout"],
          f'=IF(OR({HC["lockout_end_date"]}{r}="",NOT(ISNUMBER({HC["lockout_end_date"]}{r}))),0,'
          f'MAX(0,MIN(60,'
          f'(VALUE(LEFT({HC["period"]}{r},4))-'
          f'VALUE(LEFT(TEXT({HC["lockout_end_date"]}{r},"00000000"),4)))*12+'
          f'(VALUE(MID({HC["period"]}{r},5,2))-'
          f'VALUE(MID(TEXT({HC["lockout_end_date"]}{r},"00000000"),5,2))))))')

        # in_lockout = (period < lockout_end_yyyymm)
        f(HC["in_lockout"],
          f'=IF(OR({HC["lockout_end_date"]}{r}="",NOT(ISNUMBER({HC["lockout_end_date"]}{r}))),0,'
          f'IF(VALUE({HC["period"]}{r})<'
          f'VALUE(LEFT(TEXT({HC["lockout_end_date"]}{r},"00000000"),6)),1,0))')

        # in_prepay_penalty
        f(HC["in_prepay_penalty"],
          f'=IF(OR({HC["prepay_end_date"]}{r}="",NOT(ISNUMBER({HC["prepay_end_date"]}{r}))),0,'
          f'IF(VALUE({HC["period"]}{r})<'
          f'VALUE(LEFT(TEXT({HC["prepay_end_date"]}{r},"00000000"),6)),1,0))')

        # past_all_restrictions
        f(HC["past_all_restrictions"],
          f'=IF(AND({HC["in_lockout"]}{r}=0,{HC["in_prepay_penalty"]}{r}=0),1,0)')

        # vintage_year: prefer override; else extract from origination_date
        f(HC["vintage_year"],
          f'=IF(ISNUMBER({HC["vintage_year_input"]}{r}),{HC["vintage_year_input"]}{r},'
          f'IFERROR(VALUE(LEFT(TEXT({HC["origination_date"]}{r},"00000000"),4)),""))')

        # vintage_median_rate via XLOOKUP into Vintage_Rates
        f(HC["vintage_median_rate"],
          f'=IFERROR(XLOOKUP({HC["vintage_year"]}{r},'
          f'Vintage_Rates!$A$2:$A$1000,Vintage_Rates!$B$2:$B$1000,""),"")')

        # sato_bps: prefer override; else compute from loan_rate − vintage_median_rate (×100)
        f(HC["sato_bps"],
          f'=IF(ISNUMBER({HC["sato_bps_input"]}{r}),{HC["sato_bps_input"]}{r},'
          f'IFERROR(({HC["loan_rate_pct"]}{r}-{HC["vintage_median_rate"]}{r})*100,""))')

        # plc_rate_bps: prefer override; else lookup
        f(HC["plc_rate_bps"],
          f'=IF(ISNUMBER({HC["plc_rate_bps_input"]}{r}),{HC["plc_rate_bps_input"]}{r},'
          f'IFERROR(XLOOKUP(VALUE({HC["period"]}{r}),'
          f'PLC_Rates!$A$2:$A$1000,PLC_Rates!$B$2:$B$1000,""),""))')

        # prepay_penalty_points: prefer override; else use the standard
        # step-down rule.  Most GNMA project loans follow the 10/9/8/.../1/0
        # schedule that drops 1 pt per year, so current penalty equals
        # years remaining until prepay_end_date (capped at 10 pts).  This
        # is an approximation — irregular schedules need the override.
        f(HC["prepay_penalty_points"],
          f'=IF(ISNUMBER({HC["prepay_penalty_points_input"]}{r}),'
          f'{HC["prepay_penalty_points_input"]}{r},'
          f'IF({HC["in_prepay_penalty"]}{r}=0,0,'
          f'MAX(0,MIN(10,'
          f'((VALUE(LEFT(TEXT({HC["prepay_end_date"]}{r},"00000000"),4))-'
          f'VALUE(LEFT({HC["period"]}{r},4)))*12+'
          f'(VALUE(MID(TEXT({HC["prepay_end_date"]}{r},"00000000"),5,2))-'
          f'VALUE(MID({HC["period"]}{r},5,2))))/12))))')

        # gross_refi_incentive_bps = loan_rate*100 − PLC
        f(HC["gross_refi_incentive_bps"],
          f'=IFERROR({HC["loan_rate_pct"]}{r}*100-{HC["plc_rate_bps"]}{r},"")')

        # refi_incentive_bps = loan_rate*100 − (PLC + (1+pen)*12.5)
        f(HC["refi_incentive_bps"],
          f'=IFERROR({HC["loan_rate_pct"]}{r}*100-'
          f'({HC["plc_rate_bps"]}{r}+(1+{HC["prepay_penalty_points"]}{r})*12.5),"")')

        # upb_millions
        f(HC["upb_millions"],
          f'=IFERROR(MIN({HC["upb"]}{r}/1000000,100),"")')

        # fha_category — substring search via the FHA_Lookup table
        f(HC["fha_category"],
          f'=IFERROR(INDEX(FHA_Lookup!$B$2:$B$11,'
          f'MATCH(TRUE,ISNUMBER(SEARCH(FHA_Lookup!$A$2:$A$11,'
          f'UPPER(SUBSTITUTE({HC["fha_program_code"]}{r}," ","")))),0)),"OTHER")')

        # loan_purpose — same lookup, column C
        f(HC["loan_purpose"],
          f'=IFERROR(INDEX(FHA_Lookup!$C$2:$C$11,'
          f'MATCH(TRUE,ISNUMBER(SEARCH(FHA_Lookup!$A$2:$A$11,'
          f'UPPER(SUBSTITUTE({HC["fha_program_code"]}{r}," ","")))),0)),"OTHER")')

        # is_nc
        f(HC["is_nc"],
          f'=IF({HC["loan_purpose"]}{r}="NC",1,0)')

        # affordable status flags
        f(HC["aff_aff"],
          f'=IF(UPPER(TRIM({HC["affordable_status"]}{r}))="AFF",1,0)')
        f(HC["aff_baf"],
          f'=IF(UPPER(TRIM({HC["affordable_status"]}{r}))="BAF",1,0)')
        f(HC["aff_mkt"],
          f'=IF(UPPER(TRIM({HC["affordable_status"]}{r}))="MKT",1,0)')

        # FHA category dummies (against reference 223f)
        for cat, hdr in [
            ("221d4", "fha_221d4"), ("223a7", "fha_223a7"), ("232", "fha_232"),
            ("538", "fha_538"),     ("241", "fha_241"),     ("220", "fha_220"),
            ("OTHER", "fha_other"),
        ]:
            f(HC[hdr], f'=IF({HC["fha_category"]}{r}="{cat}",1,0)')

        # Pool-type-derived flags (replaces is_modified / is_non_level /
        # is_mature, which are 0 in 100% of the modern panel — see
        # SanCap primer; pool_type carries the modification signal).
        f(HC["is_lm_pool"],
          f'=IF(UPPER(TRIM({HC["pool_type"]}{r}))="LM",1,0)')
        f(HC["is_pn_pool"],
          f'=IF(UPPER(TRIM({HC["pool_type"]}{r}))="PN",1,0)')
        f(HC["is_ls_pool"],
          f'=IF(UPPER(TRIM({HC["pool_type"]}{r}))="LS",1,0)')
        f(HC["is_rx_pool"],
          f'=IF(UPPER(TRIM({HC["pool_type"]}{r}))="RX",1,0)')

        # ── LOG-ODDS CONTRIBUTIONS ─────────────────────────────────
        # Build piecewise-linear hinge sums for each spline group.

        def hinge_sum(value_col: str, knots: list[float],
                      lin_param: str, knot_param_prefix: str) -> str:
            terms = [f"{coef_cell(lin_param)}*{value_col}{r}"]
            for k in knots:
                kname = f"{knot_param_prefix}{int(k) if k == int(k) else k}"
                terms.append(
                    f"{coef_cell(kname)}*MAX({value_col}{r}-({k}),0)")
            return "(" + "+".join(terms) + ")"

        # REFI: input = refi_incentive_bps
        f(HC["REFI_logodds"],
          "=IFERROR(" + hinge_sum(HC["refi_incentive_bps"],
                                   META["knots"]["refi"],
                                   "refi_lin", "refi_k") + ",0)")

        # AGE
        f(HC["AGE_logodds"],
          "=IFERROR(" + hinge_sum(HC["loan_age_months"],
                                   META["knots"]["age"],
                                   "age_lin", "age_k") + ",0)")

        # PEN — pen knots are 0,2,5,8 (and pen_lin)
        f(HC["PEN_logodds"],
          "=IFERROR(" + hinge_sum(HC["prepay_penalty_points"],
                                   META["knots"]["pen"],
                                   "pen_lin", "pen_k") + ",0)")

        # SATO
        f(HC["SATO_logodds"],
          "=IFERROR(" + hinge_sum(HC["sato_bps"],
                                   META["knots"]["sato"],
                                   "sato_lin", "sato_k") + ",0)")

        # SIZE
        f(HC["SIZE_logodds"],
          "=IFERROR(" + hinge_sum(HC["upb_millions"],
                                   META["knots"]["size"],
                                   "size_lin", "size_k") + ",0)")

        # M2M
        f(HC["M2M_logodds"],
          "=IFERROR(" + hinge_sum(HC["months_to_maturity"],
                                   META["knots"]["m2m"],
                                   "m2m_lin", "m2m_k") + ",0)")

        # MPL
        f(HC["MPL_logodds"],
          "=IFERROR(" + hinge_sum(HC["months_post_lockout"],
                                   META["knots"]["mpl"],
                                   "mpl_lin", "mpl_k") + ",0)")

        # INTERACT — refi_x_pen
        f(HC["INTERACT_logodds"],
          f'=IFERROR({coef_cell("refi_x_pen")}*'
          f'(MAX({HC["refi_incentive_bps"]}{r},0)*'
          f'MIN(MAX({HC["prepay_penalty_points"]}{r},0),10))/100,0)')

        # PHASE — in_penalty
        f(HC["PHASE_logodds"],
          f'={coef_cell("in_penalty")}*{HC["in_prepay_penalty"]}{r}')

        # FHA dummies sum
        fha_terms = " + ".join(
            f"{coef_cell(name)}*{HC[name]}{r}"
            for name in ("fha_221d4", "fha_223a7", "fha_232",
                         "fha_538", "fha_241", "fha_220", "fha_other"))
        f(HC["FHA_logodds"], f'={fha_terms}')

        # PURPOSE
        f(HC["PURPOSE_logodds"], f'={coef_cell("is_nc")}*{HC["is_nc"]}{r}')

        # AFFORDABLE
        f(HC["AFF_logodds"],
          f'={coef_cell("aff_aff")}*{HC["aff_aff"]}{r}'
          f'+{coef_cell("aff_baf")}*{HC["aff_baf"]}{r}'
          f'+{coef_cell("aff_mkt")}*{HC["aff_mkt"]}{r}')

        # POOL (pool-type-derived contributions)
        f(HC["POOL_logodds"],
          f'={coef_cell("is_lm_pool")}*{HC["is_lm_pool"]}{r}'
          f'+{coef_cell("is_pn_pool")}*{HC["is_pn_pool"]}{r}'
          f'+{coef_cell("is_ls_pool")}*{HC["is_ls_pool"]}{r}'
          f'+{coef_cell("is_rx_pool")}*{HC["is_rx_pool"]}{r}')

        # ── PREDICTION ─────────────────────────────────────────────
        f(HC["intercept_logodds"], "=Parameters!$B$2")

        # total_logodds = intercept + sum of all components
        comp_cells = " + ".join(f"{HC[c]}{r}" for c in [
            "REFI_logodds", "AGE_logodds", "PEN_logodds", "SATO_logodds",
            "SIZE_logodds", "M2M_logodds", "MPL_logodds", "INTERACT_logodds",
            "PHASE_logodds", "FHA_logodds", "PURPOSE_logodds",
            "AFF_logodds", "POOL_logodds"])
        f(HC["total_logodds"],
          f'={HC["intercept_logodds"]}{r}+{comp_cells}')

        # SMM = sigmoid(total_logodds), gated by in_lockout=0 (lockout
        # forces voluntary prepay to 0).  Optional CPR scalar applied at
        # SMM-equivalent level (so the resulting CPR is scaled).
        f(HC["predicted_SMM"],
          f'=IF({HC["in_lockout"]}{r}=1,0,'
          f'1/(1+EXP(-{HC["total_logodds"]}{r}))*Parameters!$B$3)')

        # CPR = 1 - (1-SMM)^12, expressed as a percentage
        f(HC["predicted_CPR_pct"],
          f'=(1-(1-{HC["predicted_SMM"]}{r})^12)*100')

        # Baseline CPR = sigmoid(intercept), lockout-gated so a locked
        # loan contributes 0 to the weighted baseline — matching the
        # same gate already on predicted_SMM and every CPR_attr_*.
        f(HC["baseline_CPR_pct"],
          f'=IF({HC["in_lockout"]}{r}=1,0,'
          f'(1-(1-1/(1+EXP(-Parameters!$B$2)))^12)*100)')

        # Component attribution: how many CPR percentage points the
        # component is contributing on the margin.  We take the
        # difference between full predicted CPR and the predicted CPR
        # if THIS component's log-odds were 0.  Result is signed.
        for comp_name, comp_col in [
            ("REFI", HC["REFI_logodds"]),
            ("AGE",  HC["AGE_logodds"]),
            ("PEN",  HC["PEN_logodds"]),
            ("SATO", HC["SATO_logodds"]),
            ("SIZE", HC["SIZE_logodds"]),
            ("M2M",  HC["M2M_logodds"]),
            ("MPL",  HC["MPL_logodds"]),
            ("INTERACT", HC["INTERACT_logodds"]),
            ("PHASE", HC["PHASE_logodds"]),
            ("FHA",  HC["FHA_logodds"]),
            ("PURPOSE", HC["PURPOSE_logodds"]),
            ("AFF",  HC["AFF_logodds"]),
            ("POOL", HC["POOL_logodds"]),
        ]:
            attr_col = HC[f"CPR_attr_{comp_name}"]
            f(attr_col,
              f'=IF({HC["in_lockout"]}{r}=1,0,'
              f'{HC["predicted_CPR_pct"]}{r}-'
              f'(1-(1-(1/(1+EXP(-({HC["total_logodds"]}{r}-{comp_col}{r})))*'
              f'Parameters!$B$3))^12)*100)')

    # Apply input-cell highlighting to the user-input columns in data rows
    for col_idx, (h, group, _) in enumerate(layout, start=1):
        if group == "USER":
            for r in range(3, last_row + 1):
                ws.cell(row=r, column=col_idx).fill = INPUT_FILL
        elif group == "DERIVED":
            for r in range(3, last_row + 1):
                ws.cell(row=r, column=col_idx).fill = DERIVED_FILL
        elif group == "LOGODDS":
            for r in range(3, last_row + 1):
                ws.cell(row=r, column=col_idx).fill = ATTR_FILL
        elif group == "PRED":
            for r in range(3, last_row + 1):
                ws.cell(row=r, column=col_idx).fill = TOTAL_FILL

    # Number formats
    for r in range(3, last_row + 1):
        ws.cell(row=r, column=ws[HC["loan_rate_pct"] + "2"].column).number_format = "0.000"
        ws.cell(row=r, column=ws[HC["upb"] + "2"].column).number_format = "#,##0"
        ws.cell(row=r, column=ws[HC["upb_millions"] + "2"].column).number_format = "0.000"
        ws.cell(row=r, column=ws[HC["predicted_SMM"] + "2"].column).number_format = "0.00000%"
        ws.cell(row=r, column=ws[HC["predicted_CPR_pct"] + "2"].column).number_format = "0.000"
        ws.cell(row=r, column=ws[HC["baseline_CPR_pct"] + "2"].column).number_format = "0.000"
        for c in components:
            ws.cell(row=r, column=ws[HC[c] + "2"].column).number_format = "0.0000"
        for nm in [n for n, _ in pred_cols if n.startswith("CPR_attr_")]:
            ws.cell(row=r, column=ws[HC[nm] + "2"].column).number_format = "0.000"

    autosize(ws, min_w=10, max_w=22)
    ws.freeze_panes = "G3"
    return n_rows


def coef_cell(feature_name: str) -> str:
    """Return the absolute reference to a coefficient on the Parameters
    sheet.  Coefficients live in column C starting at row 6 in the
    order they appear in COEF_NO_INT.
    """
    idx = COEF_NO_INT.index[COEF_NO_INT["feature"] == feature_name]
    if len(idx) == 0:
        raise KeyError(feature_name)
    row = int(idx[0]) + 6   # row 6 is the first coef row
    return f"Parameters!$C${row}"


# ─── Deal_Summary sheet ─────────────────────────────────────────────

def write_deal_summary(ws, loan_ws, n_loans):
    """Aggregate Loan_Input by deal_id, weighted by current UPB.

    Uses dynamic-array UNIQUE() / SUMIFS formulas so the user can paste
    in a different deal mix and the summary auto-updates.
    """
    # Find the column letters on Loan_Input we need
    def loan_col(header: str) -> str:
        for col_idx, cell in enumerate(loan_ws[2], start=1):
            if cell.value == header:
                return get_column_letter(col_idx)
        raise KeyError(header)

    # Header row
    headers = [
        "deal_id", "n_loans", "total_upb_$mn",
        "weighted_predicted_CPR_pct", "weighted_baseline_CPR_pct",
        "weighted_loan_age_months", "weighted_loan_rate_pct",
        "weighted_refi_incentive_bps", "weighted_prepay_penalty_points",
        "weighted_sato_bps", "weighted_upb_millions",
        "pct_in_lockout", "pct_in_prepay_penalty", "pct_past_all_restrictions",
        # Weighted CPR attributions
        "wtd_attr_REFI", "wtd_attr_AGE", "wtd_attr_PEN", "wtd_attr_SATO",
        "wtd_attr_SIZE", "wtd_attr_M2M", "wtd_attr_MPL", "wtd_attr_INTERACT",
        "wtd_attr_PHASE", "wtd_attr_FHA", "wtd_attr_PURPOSE",
        "wtd_attr_AFF", "wtd_attr_POOL",
        # Composition
        "pct_NC", "pct_232", "pct_221d4", "pct_223f", "pct_223a7",
        "pct_538", "pct_AFF_or_BAF",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # First column — unique deal IDs
    deal_col = loan_col("deal_id")
    rng_deal = f"Loan_Input!${deal_col}$3:${deal_col}${n_loans + 2}"

    # Use UNIQUE() to populate unique deals starting in row 2
    ws["A2"] = f"=IFERROR(UNIQUE(FILTER({rng_deal},{rng_deal}<>\"\")),\"\")"

    # For each subsequent column, build SUMIFS / SUMPRODUCT-type formula
    # that depends on the deal_id in column A of the current row.
    # We pre-fill rows 2..2+max_deals (assume up to 50 deals).
    MAX_DEALS = 50

    upb_col = loan_col("upb")
    pred_cpr_col = loan_col("predicted_CPR_pct")
    base_cpr_col = loan_col("baseline_CPR_pct")
    age_col      = loan_col("loan_age_months")
    rate_col     = loan_col("loan_rate_pct")
    refi_col     = loan_col("refi_incentive_bps")
    pen_col      = loan_col("prepay_penalty_points")
    sato_col     = loan_col("sato_bps")
    upbm_col     = loan_col("upb_millions")
    inlk_col     = loan_col("in_lockout")
    inpe_col     = loan_col("in_prepay_penalty")
    past_col     = loan_col("past_all_restrictions")
    isnc_col     = loan_col("is_nc")
    f232_col     = loan_col("fha_232")
    f221_col     = loan_col("fha_221d4")
    f223a7_col   = loan_col("fha_223a7")
    f538_col     = loan_col("fha_538")
    aff_col      = loan_col("aff_aff")
    baf_col      = loan_col("aff_baf")
    fha_cat_col  = loan_col("fha_category")

    rng_upb       = f"Loan_Input!${upb_col}$3:${upb_col}${n_loans + 2}"
    rng_pred_cpr  = f"Loan_Input!${pred_cpr_col}$3:${pred_cpr_col}${n_loans + 2}"
    rng_base_cpr  = f"Loan_Input!${base_cpr_col}$3:${base_cpr_col}${n_loans + 2}"
    rng_age       = f"Loan_Input!${age_col}$3:${age_col}${n_loans + 2}"
    rng_rate      = f"Loan_Input!${rate_col}$3:${rate_col}${n_loans + 2}"
    rng_refi      = f"Loan_Input!${refi_col}$3:${refi_col}${n_loans + 2}"
    rng_pen       = f"Loan_Input!${pen_col}$3:${pen_col}${n_loans + 2}"
    rng_sato      = f"Loan_Input!${sato_col}$3:${sato_col}${n_loans + 2}"
    rng_upbm      = f"Loan_Input!${upbm_col}$3:${upbm_col}${n_loans + 2}"
    rng_inlk      = f"Loan_Input!${inlk_col}$3:${inlk_col}${n_loans + 2}"
    rng_inpe      = f"Loan_Input!${inpe_col}$3:${inpe_col}${n_loans + 2}"
    rng_past      = f"Loan_Input!${past_col}$3:${past_col}${n_loans + 2}"
    rng_isnc      = f"Loan_Input!${isnc_col}$3:${isnc_col}${n_loans + 2}"
    rng_f232      = f"Loan_Input!${f232_col}$3:${f232_col}${n_loans + 2}"
    rng_f221      = f"Loan_Input!${f221_col}$3:${f221_col}${n_loans + 2}"
    rng_f223a7    = f"Loan_Input!${f223a7_col}$3:${f223a7_col}${n_loans + 2}"
    rng_f538      = f"Loan_Input!${f538_col}$3:${f538_col}${n_loans + 2}"
    rng_aff       = f"Loan_Input!${aff_col}$3:${aff_col}${n_loans + 2}"
    rng_baf       = f"Loan_Input!${baf_col}$3:${baf_col}${n_loans + 2}"
    rng_fha_cat   = f"Loan_Input!${fha_cat_col}$3:${fha_cat_col}${n_loans + 2}"

    attr_cols = {
        comp: loan_col(f"CPR_attr_{comp}")
        for comp in ["REFI", "AGE", "PEN", "SATO", "SIZE", "M2M", "MPL",
                     "INTERACT", "PHASE", "FHA", "PURPOSE", "AFF", "POOL"]
    }
    rng_attr = {
        comp: f"Loan_Input!${col}$3:${col}${n_loans + 2}"
        for comp, col in attr_cols.items()
    }

    for r in range(2, 2 + MAX_DEALS):
        deal_ref = f"$A{r}"

        # n_loans = COUNTIF for this deal
        ws.cell(row=r, column=2,
                value=f'=IF({deal_ref}="","",COUNTIF({rng_deal},{deal_ref}))')

        # total_upb / 1e6
        ws.cell(row=r, column=3,
                value=f'=IF({deal_ref}="","",SUMIF({rng_deal},{deal_ref},{rng_upb})/1000000)')

        # weighted predicted_CPR
        ws.cell(row=r, column=4,
                value=f'=IF({deal_ref}="","",'
                      f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*{rng_pred_cpr}*{rng_upb})/'
                      f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')

        # weighted baseline CPR (just intercept-only)
        ws.cell(row=r, column=5,
                value=f'=IF({deal_ref}="","",'
                      f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*{rng_base_cpr}*{rng_upb})/'
                      f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')

        # weighted features
        for col, rng in [
            (6,  rng_age),
            (7,  rng_rate),
            (8,  rng_refi),
            (9,  rng_pen),
            (10, rng_sato),
            (11, rng_upbm),
        ]:
            ws.cell(row=r, column=col,
                    value=f'=IF({deal_ref}="","",'
                          f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*{rng}*{rng_upb})/'
                          f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')

        # composition flags (UPB-weighted)
        for col, rng in [
            (12, rng_inlk),
            (13, rng_inpe),
            (14, rng_past),
        ]:
            ws.cell(row=r, column=col,
                    value=f'=IF({deal_ref}="","",'
                          f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*{rng}*{rng_upb})/'
                          f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')

        # weighted attribution columns
        col_idx = 15
        for comp in ["REFI", "AGE", "PEN", "SATO", "SIZE", "M2M", "MPL",
                     "INTERACT", "PHASE", "FHA", "PURPOSE", "AFF", "POOL"]:
            rng = rng_attr[comp]
            ws.cell(row=r, column=col_idx,
                    value=f'=IF({deal_ref}="","",'
                          f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*{rng}*{rng_upb})/'
                          f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')
            col_idx += 1

        # Composition: pct_NC, pct_232, pct_221d4, pct_223f, pct_223a7,
        #              pct_538, pct_AFF_or_BAF
        for col, rng in [
            (col_idx,     rng_isnc),
            (col_idx + 1, rng_f232),
            (col_idx + 2, rng_f221),
        ]:
            ws.cell(row=r, column=col,
                    value=f'=IF({deal_ref}="","",'
                          f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*{rng}*{rng_upb})/'
                          f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')
        # pct_223f via fha_category lookup
        ws.cell(row=r, column=col_idx + 3,
                value=f'=IF({deal_ref}="","",'
                      f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*({rng_fha_cat}="223f")*{rng_upb})/'
                      f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')

        ws.cell(row=r, column=col_idx + 4,
                value=f'=IF({deal_ref}="","",'
                      f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*{rng_f223a7}*{rng_upb})/'
                      f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')
        ws.cell(row=r, column=col_idx + 5,
                value=f'=IF({deal_ref}="","",'
                      f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*{rng_f538}*{rng_upb})/'
                      f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')
        # pct_AFF_or_BAF
        ws.cell(row=r, column=col_idx + 6,
                value=f'=IF({deal_ref}="","",'
                      f'IFERROR(SUMPRODUCT(({rng_deal}={deal_ref})*({rng_aff}+{rng_baf})*{rng_upb})/'
                      f'SUMIF({rng_deal},{deal_ref},{rng_upb}),""))')

    # Number formats
    for r in range(2, 2 + MAX_DEALS):
        ws.cell(row=r, column=3).number_format  = "#,##0.0"      # UPB $mn
        ws.cell(row=r, column=4).number_format  = "0.00"         # CPR
        ws.cell(row=r, column=5).number_format  = "0.00"         # baseline CPR
        ws.cell(row=r, column=6).number_format  = "0.0"          # age
        ws.cell(row=r, column=7).number_format  = "0.000"        # rate
        ws.cell(row=r, column=8).number_format  = "0"            # refi inc
        ws.cell(row=r, column=9).number_format  = "0.00"         # pen
        ws.cell(row=r, column=10).number_format = "0"            # sato
        ws.cell(row=r, column=11).number_format = "0.00"         # upb_m
        for c in range(12, 15):
            ws.cell(row=r, column=c).number_format = "0.0%"
        for c in range(15, 28):
            ws.cell(row=r, column=c).number_format = "0.00"
        for c in range(28, 35):
            ws.cell(row=r, column=c).number_format = "0.0%"

    autosize(ws, min_w=10, max_w=20)
    ws.freeze_panes = "B2"


# ─── Deal_Comparison sheet ──────────────────────────────────────────

def write_deal_comparison(ws):
    """Side-by-side display for two user-selected deals.

    The two deal_id's go in cells B1 and C1.  All metrics below pull
    from Deal_Summary via VLOOKUP/INDEX.
    """
    ws["A1"] = "Metric"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = HEADER_FONT
    ws["B1"] = "Deal A → enter deal_id"
    ws["B1"].fill = INPUT_FILL
    ws["B1"].font = BOLD
    ws["C1"] = "Deal B → enter deal_id"
    ws["C1"].fill = INPUT_FILL
    ws["C1"].font = BOLD
    ws["D1"] = "Difference (A − B)"
    ws["D1"].fill = HEADER_FILL
    ws["D1"].font = HEADER_FONT
    ws["E1"] = "Driver"
    ws["E1"].fill = HEADER_FILL
    ws["E1"].font = HEADER_FONT

    # Pre-fill the first two unique deals as defaults
    deals = SAMPLE["deal_id"].drop_duplicates().tolist()
    if len(deals) >= 2:
        ws["B1"] = deals[0]
        ws["C1"] = deals[1]
    elif len(deals) >= 1:
        ws["B1"] = deals[0]

    # Column index map for Deal_Summary (matches headers list)
    summary_cols = [
        # (display_label, deal_summary_column_index_1based, fmt, driver_text)
        ("# of loans",                   2,  "#,##0",   ""),
        ("Total UPB ($mn)",              3,  "#,##0.0", ""),
        ("Weighted predicted CPR (%)",   4,  "0.00",    "(top-line model output)"),
        ("Weighted baseline CPR (%)",    5,  "0.00",    "(intercept only — model floor)"),
        ("Weighted loan age (months)",   6,  "0.0",     "Drives age/burnout effect"),
        ("Weighted loan rate (%)",       7,  "0.000",   ""),
        ("Weighted refi incentive (bps)", 8, "0",       "Primary refi S-curve driver"),
        ("Weighted penalty pts",         9,  "0.00",    "Dampens refi response"),
        ("Weighted SATO (bps)",         10,  "0",       "Borrower quality"),
        ("Weighted UPB ($mn per loan)", 11,  "0.00",    "Larger loans refi more"),
        ("% in lockout (UPB-wtd)",      12,  "0.0%",    "Hard-zero floor"),
        ("% in penalty window",         13,  "0.0%",    "Penalty applies"),
        ("% open window",               14,  "0.0%",    "No call protection"),
        # CPR attributions
        ("CPR attr: REFI",              15,  "0.00",    "Refi incentive"),
        ("CPR attr: AGE",               16,  "0.00",    "Loan age / seasoning / burnout"),
        ("CPR attr: PENALTY",           17,  "0.00",    "Penalty point S-curve"),
        ("CPR attr: SATO",              18,  "0.00",    "Spread at origination"),
        ("CPR attr: SIZE",              19,  "0.00",    "Loan size effect"),
        ("CPR attr: M2M",               20,  "0.00",    "Months to maturity"),
        ("CPR attr: MPL",               21,  "0.00",    "Months post lockout"),
        ("CPR attr: INTERACT",          22,  "0.00",    "Refi × penalty interaction"),
        ("CPR attr: PHASE",             23,  "0.00",    "In-penalty phase indicator"),
        ("CPR attr: FHA",               24,  "0.00",    "FHA program (vs 223f)"),
        ("CPR attr: PURPOSE",           25,  "0.00",    "NC vs RP"),
        ("CPR attr: AFFORDABLE",        26,  "0.00",    "AFF/BAF/MKT vs unknown"),
        ("CPR attr: POOL",              27,  "0.00",    "Pool type (LM/PN/LS/RX dummies)"),
        # Composition
        ("Composition: % NC",           28,  "0.0%",    ""),
        ("Composition: % 232 healthcare", 29,  "0.0%",  ""),
        ("Composition: % 221(d)(4)",    30,  "0.0%",    ""),
        ("Composition: % 223(f)",       31,  "0.0%",    ""),
        ("Composition: % 223(a)(7)",    32,  "0.0%",    ""),
        ("Composition: % 538 USDA",     33,  "0.0%",    ""),
        ("Composition: % AFF or BAF",   34,  "0.0%",    ""),
    ]

    for i, (label, col_idx, fmt, driver_text) in enumerate(summary_cols, start=2):
        ws.cell(row=i, column=1, value=label).font = BOLD

        col_letter_summary = get_column_letter(col_idx)
        # Lookup deal A
        ws.cell(row=i, column=2,
                value=f'=IFERROR(VLOOKUP($B$1,Deal_Summary!$A:${col_letter_summary},'
                      f'{col_idx},FALSE),"")')
        ws.cell(row=i, column=3,
                value=f'=IFERROR(VLOOKUP($C$1,Deal_Summary!$A:${col_letter_summary},'
                      f'{col_idx},FALSE),"")')

        # Difference (A - B); only meaningful for numeric cells
        ws.cell(row=i, column=4,
                value=f'=IFERROR(B{i}-C{i},"")')
        ws.cell(row=i, column=5, value=driver_text)

        ws.cell(row=i, column=2).number_format = fmt
        ws.cell(row=i, column=3).number_format = fmt
        ws.cell(row=i, column=4).number_format = fmt

        # Section coloring
        if "CPR attr:" in label:
            for c in (1, 2, 3, 4, 5):
                ws.cell(row=i, column=c).fill = ATTR_FILL
        elif label.startswith("Composition"):
            for c in (1, 2, 3, 4, 5):
                ws.cell(row=i, column=c).fill = SECTION_FILL
        else:
            for c in (1, 2, 3, 4, 5):
                ws.cell(row=i, column=c).fill = DERIVED_FILL

    autosize(ws, min_w=14, max_w=40)


# ─── Validation sheet ───────────────────────────────────────────────

def write_validation(ws):
    """Static snapshot of train/test metrics + decile calibration table."""
    ws["A1"] = "Model validation summary"
    ws["A1"].font = Font(bold=True, size=14)

    snap = [
        ("Train period", f'{META["train_period_min"]} – {META["train_period_max"]}'),
        ("Train observations", META["train_n"]),
        ("Train events (voluntary prepays)", META["train_events"]),
        ("Train AUC", round(META["train_auc"], 4)),
        ("Train predicted CPR (%)", round(META["train_pred_cpr_pct"], 2)),
        ("Train actual CPR (%)", round(META["train_actual_cpr_pct"], 2)),
        ("", ""),
        ("Test period", f'{META["test_period_min"]} – {META["test_period_max"]}'),
        ("Test observations", META["test_n"]),
        ("Test events (voluntary prepays)", META["test_events"]),
        ("Test AUC", round(META["test_auc"], 4)),
        ("Test predicted CPR (%)", round(META["test_pred_cpr_pct"], 2)),
        ("Test actual CPR (%)", round(META["test_actual_cpr_pct"], 2)),
    ]
    for i, (k, v) in enumerate(snap, start=3):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)

    # Calibration deciles
    cal = pd.read_csv(os.path.join(MODEL_DIR, "calibration.csv"))
    start = len(snap) + 5
    ws.cell(row=start, column=1, value="Out-of-sample calibration by predicted decile").font = Font(bold=True, size=12)
    headers = list(cal.columns)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start + 1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, row in cal.iterrows():
        for j, h in enumerate(headers, start=1):
            ws.cell(row=start + 2 + i, column=j, value=row[h])

    # Segment validation
    seg = pd.read_csv(os.path.join(MODEL_DIR, "segment_validation.csv"))
    start2 = start + len(cal) + 5
    ws.cell(row=start2, column=1, value="Out-of-sample CPR by segment").font = Font(bold=True, size=12)
    headers = list(seg.columns)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start2 + 1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, row in seg.iterrows():
        for j, h in enumerate(headers, start=1):
            ws.cell(row=start2 + 2 + i, column=j, value=row[h])

    autosize(ws)


# ─── Named ranges ───────────────────────────────────────────────────

def define_named_ranges(wb):
    # Just a couple for readability — coefficients are referenced
    # directly via Parameters!$C$6 etc.  Could expand later.
    pass


if __name__ == "__main__":
    build()
